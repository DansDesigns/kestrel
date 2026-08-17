"""Turning a dropped file into something a model can read.

A local model reads text. Everything here exists to get text out of the formats
people actually have — a Word document, a spreadsheet, a PDF, a screenshot —
without requiring a stack of libraries to be installed first. Where a proper
library is available it is used; where it is not, the fallback opens the file
and takes what it can, because a rough extraction is far better than "cannot
read this".

Images are the exception: their content cannot be read without a vision model,
so what is recorded is what can be known — format, dimensions, size — and the
person is told plainly rather than being left to wonder why the model ignored
the picture.
"""
from __future__ import annotations

import json
import re
import struct
import zipfile
from dataclasses import dataclass, field
from pathlib import Path

TEXT_SUFFIXES = {
    ".txt", ".md", ".markdown", ".rst", ".log", ".csv", ".tsv", ".json", ".yaml",
    ".yml", ".toml", ".ini", ".cfg", ".conf", ".env", ".py", ".js", ".ts", ".jsx",
    ".tsx", ".rs", ".go", ".c", ".h", ".cpp", ".hpp", ".java", ".kt", ".rb",
    ".php", ".sh", ".bat", ".ps1", ".sql", ".html", ".htm", ".css", ".xml",
    ".svg", ".gradle", ".make", ".cmake", ".diff", ".patch",
}
IMAGE_SUFFIXES = {".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp", ".tiff",
                  ".ico", ".heic"}
OFFICE_SUFFIXES = {".docx", ".xlsx", ".pptx", ".odt", ".ods", ".odp"}

MAX_CHARS = 40_000          # a whole book helps nobody inside a 4k window


@dataclass
class Attachment:
    path: Path
    name: str = ""
    kind: str = "text"       # text | image | document | binary
    text: str = ""
    note: str = ""           # what could not be read, in plain words
    size: int = 0
    truncated: bool = False
    meta: dict = field(default_factory=dict)

    @property
    def label(self) -> str:
        from .models import human_size
        bits = [self.name or self.path.name, human_size(self.size)]
        if self.kind == "image" and self.meta.get("dimensions"):
            bits.append(self.meta["dimensions"])
        elif self.text:
            bits.append(f"{len(self.text.splitlines())} lines")
        return " · ".join(bits)

    def block(self) -> str:
        """How it is put to the model."""
        if self.kind == "image":
            return (f"[image: {self.name}, {self.meta.get('dimensions', 'size unknown')}. "
                    "Its contents cannot be read by a text model.]")
        if not self.text.strip():
            return f"[{self.name}: {self.note or 'no readable text'}]"
        cut = "\n[… truncated]" if self.truncated else ""
        return f"--- {self.name} ---\n{self.text}{cut}\n--- end of {self.name} ---"


def read(path: str | Path, limit: int = MAX_CHARS) -> Attachment:
    p = Path(path).expanduser()
    item = Attachment(path=p, name=p.name)
    try:
        item.size = p.stat().st_size
    except OSError as e:
        item.note = str(e)
        item.kind = "binary"
        return item

    suffix = p.suffix.lower()
    if suffix in IMAGE_SUFFIXES:
        item.kind = "image"
        item.meta["dimensions"] = _image_size(p)
        item.note = "images need a vision model to be read"
        return item
    if suffix in OFFICE_SUFFIXES:
        item.kind = "document"
        item.text, item.note = _office_text(p, suffix)
    elif suffix == ".pdf":
        item.kind = "document"
        item.text, item.note = _pdf_text(p)
    elif suffix in TEXT_SUFFIXES or item.size < 2_000_000:
        item.kind = "text"
        item.text, item.note = _plain_text(p)
    else:
        item.kind = "binary"
        item.note = "not a readable text format"

    if len(item.text) > limit:
        item.text = item.text[:limit]
        item.truncated = True
    return item


# -- plain files -------------------------------------------------------------
def _plain_text(p: Path) -> tuple[str, str]:
    try:
        raw = p.read_bytes()
    except OSError as e:
        return "", str(e)
    if b"\x00" in raw[:2048]:
        return "", "looks like a binary file"
    for encoding in ("utf-8", "utf-16", "cp1252", "latin-1"):
        try:
            return raw.decode(encoding), ""
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace"), "some characters were replaced"


# -- office ------------------------------------------------------------------
def _office_text(p: Path, suffix: str) -> tuple[str, str]:
    """Read the modern formats without needing a library installed.

    A .docx or .xlsx is a zip of XML. Stripping the tags gives the words in
    order, which is what a model needs — the formatting is not the point.
    """
    if suffix == ".docx":
        try:
            import docx                                   # noqa: F401
            from docx import Document
            return "\n".join(x.text for x in Document(str(p)).paragraphs), ""
        except Exception:
            # Fall through to the zip reader: a library that cannot open a file
            # is not a reason to give up when the format is a zip of XML.
            pass
    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
            book = load_workbook(str(p), read_only=True, data_only=True)
            out = []
            for sheet in book.worksheets:
                out.append(f"# {sheet.title}")
                for row in sheet.iter_rows(values_only=True):
                    if any(c is not None for c in row):
                        out.append("\t".join("" if c is None else str(c) for c in row))
            return "\n".join(out), ""
        except Exception:
            pass

    inner = {".docx": ["word/document.xml"],
             ".pptx": None,                                 # every slide
             ".xlsx": ["xl/sharedStrings.xml"],
             ".odt": ["content.xml"], ".ods": ["content.xml"],
             ".odp": ["content.xml"]}.get(suffix)
    try:
        with zipfile.ZipFile(p) as archive:
            names = (inner if inner is not None else
                     [n for n in archive.namelist()
                      if n.startswith("ppt/slides/slide") and n.endswith(".xml")])
            chunks = []
            for name in names or []:
                try:
                    chunks.append(archive.read(name).decode("utf-8", "replace"))
                except KeyError:
                    continue
            if not chunks:
                return "", "no readable text inside"
            return _strip_xml("\n".join(chunks)), "read without formatting"
    except (zipfile.BadZipFile, OSError) as e:
        return "", f"could not open: {e}"


def _strip_xml(xml: str) -> str:
    # Paragraph and row ends become line breaks before the tags are removed,
    # or the whole document arrives as one unreadable run.
    xml = re.sub(r"</(w:p|a:p|text:p|row)>", "\n", xml)
    xml = re.sub(r"<[^>]+>", " ", xml)
    xml = xml.replace("&amp;", "&").replace("&lt;", "<").replace("&gt;", ">")
    lines = [" ".join(line.split()) for line in xml.splitlines()]
    return "\n".join(line for line in lines if line)


# -- pdf ---------------------------------------------------------------------
def _pdf_text(p: Path) -> tuple[str, str]:
    try:
        from pypdf import PdfReader
    except ImportError:
        try:
            from PyPDF2 import PdfReader           # older name
        except ImportError:
            return "", ("PDF text needs pypdf — pip install pypdf, or paste the "
                        "text in instead")
    try:
        reader = PdfReader(str(p))
        pages = [page.extract_text() or "" for page in reader.pages[:80]]
        text = "\n\n".join(pages).strip()
        if not text:
            return "", "this PDF has no text layer — it may be scanned images"
        return text, ""
    except Exception as e:
        return "", f"could not read the PDF: {e}"


# -- images ------------------------------------------------------------------
def _image_size(p: Path) -> str:
    """Dimensions without requiring an imaging library."""
    try:
        with open(p, "rb") as handle:
            head = handle.read(32)
            if head[:8] == b"\x89PNG\r\n\x1a\n":
                w, h = struct.unpack(">II", head[16:24])
                return f"{w}x{h}"
            if head[:2] == b"\xff\xd8":             # jpeg: walk the segments
                handle.seek(2)
                while True:
                    marker = handle.read(2)
                    if len(marker) < 2 or marker[0] != 0xFF:
                        break
                    length = struct.unpack(">H", handle.read(2))[0]
                    if 0xC0 <= marker[1] <= 0xCF and marker[1] not in (0xC4, 0xC8, 0xCC):
                        data = handle.read(5)
                        h, w = struct.unpack(">HH", data[1:5])
                        return f"{w}x{h}"
                    handle.seek(length - 2, 1)
            if head[:6] in (b"GIF87a", b"GIF89a"):
                w, h = struct.unpack("<HH", head[6:10])
                return f"{w}x{h}"
    except (OSError, struct.error):
        pass
    return "size unknown"


def summarise(items: list[Attachment]) -> str:
    """The block of attachments as it is added to a message."""
    if not items:
        return ""
    parts = [item.block() for item in items]
    return "\n\n".join(parts)


def as_json(items: list[Attachment]) -> str:
    return json.dumps([{"name": i.name, "kind": i.kind, "size": i.size}
                       for i in items])
